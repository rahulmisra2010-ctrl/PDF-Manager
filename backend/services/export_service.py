import csv
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class ExportService:
    """Export document fields to CSV, Excel, or JSON."""

    def export(self, doc, fields, fmt: str, export_folder: str) -> str:
        os.makedirs(export_folder, exist_ok=True)
        fname = f"doc_{doc.id}_{int(datetime.utcnow().timestamp())}.{fmt}"
        path = os.path.join(export_folder, fname)
        if fmt == 'csv':
            self._export_csv(doc, fields, path)
        elif fmt == 'xlsx':
            self._export_xlsx(doc, fields, path)
        elif fmt == 'json':
            self._export_json(doc, fields, path)
        return path

    def _export_csv(self, doc, fields, path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Document', 'Field Name', 'Field Value', 'Confidence', 'Page', 'Edited', 'Approved'])
            for field in fields:
                writer.writerow([
                    doc.filename,
                    field.field_name,
                    field.field_value,
                    field.confidence,
                    field.page_number,
                    field.is_edited,
                    field.is_approved
                ])

    def _export_xlsx(self, doc, fields, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Extracted Fields'
        headers = ['Document', 'Field Name', 'Field Value', 'Confidence', 'Page', 'Edited', 'Approved']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row, field in enumerate(fields, 2):
            ws.cell(row=row, column=1, value=doc.filename)
            ws.cell(row=row, column=2, value=field.field_name)
            ws.cell(row=row, column=3, value=field.field_value)
            ws.cell(row=row, column=4, value=round(field.confidence, 3))
            ws.cell(row=row, column=5, value=field.page_number)
            ws.cell(row=row, column=6, value='Yes' if field.is_edited else 'No')
            ws.cell(row=row, column=7, value='Yes' if field.is_approved else 'No')
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        wb.save(path)

    def _export_json(self, doc, fields, path):
        data = {
            'document': {
                'id': doc.id,
                'filename': doc.filename,
                'status': doc.status,
                'page_count': doc.page_count,
            },
            'fields': [
                {
                    'field_name': f.field_name,
                    'field_value': f.field_value,
                    'confidence': f.confidence,
                    'page_number': f.page_number,
                    'is_edited': f.is_edited,
                    'is_approved': f.is_approved,
                }
                for f in fields
            ]
        }
        with open(path, 'w', encoding='utf-8') as fp:
            json.dump(data, fp, indent=2, default=str)
